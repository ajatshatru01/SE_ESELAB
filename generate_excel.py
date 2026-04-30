"""
Generate test_cases.xlsx with two sheets:
  Sheet 1 – Black Box Test Cases (15 rows)
  Sheet 2 – White Box Test Cases (15 rows)

Each row: ID | Description | Input | Expected Output

Run:
    python generate_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  Colour palette
# ─────────────────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D6E4F0"
ACCENT_GREEN= "1A7A4A"
LIGHT_GREEN = "D6F0E0"
DARK_RED    = "8B0000"
LIGHT_RED   = "FAD7D7"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"
DARK_GREEN  = "1E4620"
PASTEL_GREEN= "E8F5E9"

# ─────────────────────────────────────────────────────────────────────────────
#  Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def hdr_font(colour=WHITE):
    return Font(name="Calibri", bold=True, color=colour, size=11)

def body_font(bold=False, colour="000000"):
    return Font(name="Calibri", bold=bold, color=colour, size=10)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="444444")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
#  Black Box Test Cases
# ─────────────────────────────────────────────────────────────────────────────
BLACK_BOX = [
    (
        "BB-TC-01",
        "Valid withdrawal with correct PIN and sufficient balance",
        "Card: 4111111111111111 | PIN: 1234 | Balance: Rs.10,000 | Amount: Rs.500",
        "Status: SUCCESS | Cash dispensed: Rs.500 | Remaining Balance: Rs.9,500 | Receipt generated"
    ),
    (
        "BB-TC-02",
        "Single wrong PIN entry",
        "Card: 4111111111111111 | PIN entered: 0000 (wrong)",
        "AuthenticationError raised | Card NOT blocked | 2 attempts remaining"
    ),
    (
        "BB-TC-03",
        "Card blocked after 3 consecutive wrong PINs",
        "Card: 4111111111111111 | PIN entered: 9999 (×3 wrong)",
        "CardBlockedError raised on 3rd attempt | Card status: BLOCKED"
    ),
    (
        "BB-TC-04",
        "Attempt to use an already blocked card",
        "Card: blocked card | PIN: 1234 (correct)",
        "CardBlockedError raised immediately; no authentication attempted"
    ),
    (
        "BB-TC-05",
        "Withdrawal with insufficient account balance",
        "Card: valid | PIN: 1234 | Balance: Rs.200 | Amount: Rs.500",
        "InsufficientFundsError raised | Balance unchanged: Rs.200 | No cash dispensed"
    ),
    (
        "BB-TC-06",
        "ATM has less cash than the requested withdrawal amount",
        "ATM cash: Rs.300 | Account balance: Rs.10,000 | Amount: Rs.500",
        "ATMOutOfCashError raised | Account balance unchanged | No cash dispensed"
    ),
    (
        "BB-TC-07",
        "Withdrawal amount of zero rupees",
        "Card: valid | PIN: 1234 | Amount: Rs.0",
        "InvalidAmountError raised | 'must be greater than zero'"
    ),
    (
        "BB-TC-08",
        "Negative withdrawal amount entered",
        "Card: valid | PIN: 1234 | Amount: Rs.-500",
        "InvalidAmountError raised | 'must be greater than zero'"
    ),
    (
        "BB-TC-09",
        "Amount not a multiple of Rs.100 (denomination rule)",
        "Card: valid | PIN: 1234 | Amount: Rs.150",
        "InvalidAmountError raised | 'must be a multiple of Rs.100'"
    ),
    (
        "BB-TC-10",
        "Withdrawal of exact account balance (boundary)",
        "Card: valid | PIN: 1234 | Balance: Rs.1,000 | Amount: Rs.1,000",
        "Status: SUCCESS | Balance: Rs.0 | Receipt generated"
    ),
    (
        "BB-TC-11",
        "Bank server offline during transaction",
        "Card: valid | PIN: 1234 | Balance: Rs.10,000 | Server: OFFLINE | Amount: Rs.500",
        "ServerError raised | Account balance unchanged | No cash dispensed"
    ),
    (
        "BB-TC-12",
        "Receipt is generated after a successful withdrawal",
        "Card: valid | PIN: 1234 | Balance: Rs.10,000 | Amount: Rs.500",
        "Receipt contains: Txn ID, Account (masked), Amount Rs.500, Status SUCCESS, Timestamp"
    ),
    (
        "BB-TC-13",
        "Two consecutive successful withdrawals in same session",
        "Card: valid | PIN: 1234 | Balance: Rs.5,000 | Withdrawal 1: Rs.1,000 | Withdrawal 2: Rs.2,000",
        "Both transactions SUCCESS | Final balance: Rs.2,000 | Two receipts generated"
    ),
    (
        "BB-TC-14",
        "Successful authentication after one wrong PIN attempt",
        "Card: valid | Wrong PIN once, then correct PIN: 1234",
        "Auth SUCCESS on 2nd attempt | failed_attempts counter reset to 0 | Session active"
    ),
    (
        "BB-TC-15",
        "Fractional (decimal) withdrawal amount",
        "Card: valid | PIN: 1234 | Amount: Rs.100.50",
        "InvalidAmountError raised | 'must be a whole number'"
    ),
]

# ─────────────────────────────────────────────────────────────────────────────
#  White Box Test Cases
# ─────────────────────────────────────────────────────────────────────────────
WHITE_BOX = [
    (
        "WB-TC-01",
        "Card._failed_attempts counter increments by 1 per wrong PIN (Card logic)",
        "Call card.increment_failed_attempts() twice on a fresh Card object",
        "card.failed_attempts == 2 after two increments"
    ),
    (
        "WB-TC-02",
        "Card blocked exactly at MAX_PIN_ATTEMPTS=3 (boundary condition)",
        "Call increment_failed_attempts() 2 times → check not blocked; call 3rd time",
        "is_blocked == False after 2 calls; is_blocked == True after 3rd call"
    ),
    (
        "WB-TC-03",
        "Card.reset_failed_attempts() zeroes the internal counter after failures",
        "Increment 2 times; then call reset_failed_attempts()",
        "failed_attempts == 0 after reset"
    ),
    (
        "WB-TC-04",
        "Card.validate_pin() returns False for incorrect PIN (False branch)",
        "Card(pin='1234'); call validate_pin('0000')",
        "Return value is False; no exception raised; failed_attempts unchanged"
    ),
    (
        "WB-TC-05",
        "Card.validate_pin() raises CardBlockedError when is_blocked==True (guard branch)",
        "card.block(); then call validate_pin('1234')",
        "CardBlockedError raised before PIN comparison is evaluated"
    ),
    (
        "WB-TC-06",
        "Account.debit() reduces _balance by exact amount (success path)",
        "Account(balance=10000); debit(3000)",
        "_balance == 7000; return value == 7000"
    ),
    (
        "WB-TC-07",
        "Account.debit() raises InsufficientFundsError when amount > _balance",
        "Account(balance=100); debit(200)",
        "InsufficientFundsError raised; _balance still == 100"
    ),
    (
        "WB-TC-08",
        "Account.debit() raises ValueError on zero or negative amount (guard branch)",
        "debit(0); debit(-100) on Account(balance=10000)",
        "ValueError raised for both; balance unchanged"
    ),
    (
        "WB-TC-09",
        "_validate_amount() zero-amount path (first if-branch)",
        "WithdrawController._validate_amount(0)",
        "InvalidAmountError raised; message contains 'greater than zero'"
    ),
    (
        "WB-TC-10",
        "_validate_amount() fractional amount path (second if-branch)",
        "WithdrawController._validate_amount(200.75)",
        "InvalidAmountError raised; message contains 'whole number'"
    ),
    (
        "WB-TC-11",
        "_validate_amount() non-multiple-of-100 path (third if-branch)",
        "WithdrawController._validate_amount(250)",
        "InvalidAmountError raised; message contains 'multiple of Rs.100'"
    ),
    (
        "WB-TC-12",
        "ATM cash register (self._atm_cash) decrements by exact withdrawal amount",
        "WithdrawController(atm_cash=5000); process_withdrawal(account, 2000)",
        "wc.atm_cash == 3000 after successful withdrawal"
    ),
    (
        "WB-TC-13",
        "Transaction.mark_success() sets _status=SUCCESS and stores message",
        "txn = Transaction(); txn.mark_success('All good')",
        "txn.status == SUCCESS; txn.message == 'All good'"
    ),
    (
        "WB-TC-14",
        "Transaction.mark_failed() sets _status=FAILED and stores failure reason",
        "txn = Transaction(); txn.mark_failed('Insufficient funds')",
        "txn.status == FAILED; txn.message contains 'Insufficient'"
    ),
    (
        "WB-TC-15",
        "AuthController.end_session() clears session_active and authenticated_card",
        "Authenticate successfully; then call auth.end_session()",
        "session_active == False; authenticated_card == None"
    ),
]

# ─────────────────────────────────────────────────────────────────────────────
#  Sheet builder
# ─────────────────────────────────────────────────────────────────────────────
COLUMNS = ["Test Case ID", "Description", "Input", "Expected Output"]
COL_WIDTHS = [14, 52, 58, 58]

def style_title_row(ws, title: str, theme_fill: str):
    """Merge cells A1:D1 as a big title banner."""
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value     = title
    cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=14)
    cell.fill      = fill(theme_fill)
    cell.alignment = center()
    ws.row_dimensions[1].height = 30

def style_header_row(ws, row: int, hdr_fill: str):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font      = hdr_font(WHITE)
        cell.fill      = fill(hdr_fill)
        cell.alignment = center()
        cell.border    = medium_border()
    ws.row_dimensions[row].height = 22

def build_sheet(ws, title: str, data: list, theme_dark: str, theme_mid: str,
                row_fill_a: str, row_fill_b: str, id_colour: str):
    """Populate and style a single worksheet."""
    style_title_row(ws, title, theme_dark)
    style_header_row(ws, 2, theme_mid)

    for i, (tc_id, desc, inp, expected) in enumerate(data, start=3):
        row_fill = row_fill_a if (i % 2 == 1) else row_fill_b
        values   = [tc_id, desc, inp, expected]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill   = fill(row_fill)
            cell.border = thin_border()
            if col_idx == 1:
                cell.font      = body_font(bold=True, colour=id_colour)
                cell.alignment = center()
            else:
                cell.font      = body_font()
                cell.alignment = left()
        ws.row_dimensions[i].height = 60

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header rows
    ws.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()

    # ── Sheet 1: Black Box ───────────────────────────────────────────────
    ws_bb = wb.active
    ws_bb.title = "Black Box Test Cases"
    build_sheet(
        ws_bb,
        title      = "ATM Banking System – Black Box Test Cases (Withdraw Cash Use Case)",
        data       = BLACK_BOX,
        theme_dark = DARK_BLUE,
        theme_mid  = MID_BLUE,
        row_fill_a = WHITE,
        row_fill_b = LIGHT_BLUE,
        id_colour  = MID_BLUE,
    )

    # ── Sheet 2: White Box ───────────────────────────────────────────────
    ws_wb = wb.create_sheet("White Box Test Cases")
    build_sheet(
        ws_wb,
        title      = "ATM Banking System – White Box Test Cases (Withdraw Cash Use Case)",
        data       = WHITE_BOX,
        theme_dark = DARK_GREEN,
        theme_mid  = ACCENT_GREEN,
        row_fill_a = WHITE,
        row_fill_b = PASTEL_GREEN,
        id_colour  = ACCENT_GREEN,
    )

    out = "test_cases.xlsx"
    wb.save(out)
    print(f"✔  Excel file saved: {out}")
    print(f"   Sheet 1: Black Box Test Cases  ({len(BLACK_BOX)} rows)")
    print(f"   Sheet 2: White Box Test Cases  ({len(WHITE_BOX)} rows)")

if __name__ == "__main__":
    main()
